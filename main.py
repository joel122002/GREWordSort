import nltk
import spacy
from nltk.stem import PorterStemmer, WordNetLemmatizer
from nltk.corpus import wordnet as wn
from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.cell.read_only import EmptyCell
from collections import defaultdict

# Initialize NLP tools
stemmer = PorterStemmer()
lemmatizer = WordNetLemmatizer()
nlp = spacy.load("en_core_web_sm")

def are_derivationally_related(word1, word2):
    synsets1 = wn.synsets(word1)
    related_forms = set()
    for s in synsets1:
        for lemma in s.lemmas():
            related_forms.update([d.name() for d in lemma.derivationally_related_forms()])
    return word2 in related_forms

def have_same_root(word1, word2):
    word1 = word1.lower()
    word2 = word2.lower()

    # 1. Stemming
    if stemmer.stem(word1) == stemmer.stem(word2):
        return True

    # 2. Lemmatization (NLTK)
    if lemmatizer.lemmatize(word1) == lemmatizer.lemmatize(word2):
        return True

    # 3. Lemmatization (spaCy)
    doc = nlp(f"{word1} {word2}")
    lemmas = [token.lemma_ for token in doc]
    if lemmas[0] == lemmas[1]:
        return True

    # 4. WordNet derivationally related
    if are_derivationally_related(word1, word2) or are_derivationally_related(word2, word1):
        return True

    return False

def get_frequency_dict_from_worksheet(worksheet):
    frequency_dict: dict[int, list[str]] = defaultdict(list)
    
    for word, freq in worksheet.iter_rows(min_row=2, values_only=True):  # skip header
        if word is None or freq is None:
            continue                                              # ignore blanks
    
        frequency_dict[freq].append(word)
    frequency_dict = dict(frequency_dict)
    return frequency_dict

def get_word_from_row(row):
    word_containing_cell = row[0]  # Assuming the word is in the first column
    if word_containing_cell.value is None:
        return None
    return word_containing_cell.value.strip() 

def row_contains_highly_frequent_word(row, high_frequency_words):
    for word in high_frequency_words:
        try:
            if have_same_root(word, row[0].value):
                return True
        except:
            pass 
    return False

def get_column_widths(worksheet):
    column_widths = {}
    reference_row = 2  # Assuming the second row has the data to determine column widths
    for cell in worksheet[reference_row]:
        col_letter = cell.column_letter
        width = ws.column_dimensions[col_letter].width
        column_widths[col_letter] = width
    return column_widths


MY_WORDSHEET_PATH = "GRE Word sheet.xlsx"
NEW_WORDSHEET_PATH = "GRE Word sheet - New.xlsx"
FREQUENCY_XLSX_PATH = "word_frequencies.xlsx"

wb = load_workbook(FREQUENCY_XLSX_PATH, data_only=True)
ws = wb.active # first (or only) sheet

frequency_dict = get_frequency_dict_from_worksheet(ws)


freq_sorted = sorted(frequency_dict.keys(), reverse=True) # list of frequencies sorted in descending order

buckets = [[] for _ in range(len(freq_sorted) + 1)]  # last one = “no match”

# 3. index words that belong to each frequency for O(1) membership testing
#    (set lookup is much faster than list lookup)
freq_word_sets = {f: set(freq_words) for f, freq_words in frequency_dict.items()}

wb = load_workbook(MY_WORDSHEET_PATH)
ws = wb.active # first (or only) sheet

widths = get_column_widths(ws)  # get column widths to preserve formatting later

# 4. distribute the words
for row in ws.iter_rows(min_row=2):
    placed = False
    for idx, f in enumerate(freq_sorted): 
        if row_contains_highly_frequent_word(row, freq_word_sets[f]):
            buckets[idx].append(row)
            placed = True
            break
    if not placed:                                  # fell through → no match
        buckets[-1].append(row)

flattened = [word for bucket in buckets for word in bucket]

# Create output workbook
out_wb = Workbook()
out_ws = out_wb.active
header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
# Write header with formatting
for col_idx, cell in enumerate(header_row, start=1):
    new_cell = out_ws.cell(row=1, column=col_idx, value=cell.value)
    if not isinstance(cell, EmptyCell) and cell.has_style:
        new_cell.font = copy(cell.font)
        new_cell.fill = copy(cell.fill)
        new_cell.border = copy(cell.border)
        new_cell.alignment = copy(cell.alignment)
        new_cell.number_format = cell.number_format

# Function to copy a row with formatting
def write_row(out_ws, row_index, source_row, ws):
    for col_idx, cell in enumerate(source_row, start=1):
        new_cell = out_ws.cell(row=row_index, column=col_idx, value=cell.value)
        if not isinstance(cell, EmptyCell) and cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = cell.number_format
    out_ws.row_dimensions[row_index].height = ws.row_dimensions[row_index].height

# Write list1 then list2 rows
row_index = 2
for row in flattened:
    write_row(out_ws, row_index, row, ws)
    row_index += 1

for col_letter, width in widths.items():
    ws.column_dimensions[col_letter].width = width

# Save the output
out_wb.save(NEW_WORDSHEET_PATH)
print(f"Saved output to '{NEW_WORDSHEET_PATH}'")
